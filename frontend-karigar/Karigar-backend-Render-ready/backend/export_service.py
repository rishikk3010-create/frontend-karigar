"""Generate a full-profile XLSX export with real embedded images per worker.

CSV cannot hold a viewable image (it's a plain-text format), so the "full
profile export with photos" deliverable is an .xlsx workbook instead, using
openpyxl's native image-in-cell support. Each worker gets one row; each image
field gets its own column with the actual photo(s) embedded and visible.

This intentionally does NOT try to cram every image into a single CSV cell
as base64 text — that produces a multi-hundred-MB unreadable file that breaks
in Excel/Sheets. See conversation notes for the full rationale.
"""
import base64
import io
import logging
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PILImage

logger = logging.getLogger(__name__)

THUMB_PX = 110  # square thumbnail size embedded per image, in pixels
ROW_HEIGHT_PX_TO_PT = 0.75  # Excel row-height points ≈ pixels * 0.75
COL_WIDTH_PX_TO_CHARS = 1 / 7  # rough px-to-Excel-column-width-units conversion

HEADER_FILL = PatternFill(start_color="7A2E1D", end_color="7A2E1D", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

TEXT_COLUMNS = [
    ("Worker ID", "id"),
    ("Name", "full_name"),
    ("Mobile", "phone"),
    ("DOB", "dob"),
    ("Gender", "gender"),
    ("Languages", None),  # joined list, handled specially
    ("Skills", None),
    ("City", "city"),
    ("Area", "area"),
    ("Years Experience", "years_experience"),
    ("Current Employer", "current_employer"),
    ("Previous Employer", "previous_employer"),
    ("Wage Expectation", "wage_expectation"),
    ("PhonePe/GPay Number", "upi_id"),
    ("Employment Proof Type", "employment_proof_type"),
    ("Availability Status", "availability_status"),
    ("Available From", "available_from"),
    ("Verification Status", "verification_status"),
    ("Rejection Reason", "rejection_reason"),
    ("Referral Code", "referral_code"),
    ("Referred By Code", "referred_by_code"),
    ("Duplicate Flags", None),  # joined list, handled specially
    ("Registration Date", "created_at"),
]

IMAGE_COLUMNS = [
    ("Aadhaar Card", "aadhar_images"),
    ("Employment Proof", "employment_proof_images"),
    ("Portfolio", "portfolio_images"),
]


def _decode_thumbnail(data_url: str) -> Optional[io.BytesIO]:
    """Decode a base64 data URL and downscale to a square thumbnail so the
    workbook stays a manageable size with many images per row."""
    if not data_url:
        return None
    try:
        b64 = data_url.split(",", 1)[1] if "," in data_url else data_url
        raw = base64.b64decode(b64)
        img = PILImage.open(io.BytesIO(raw))
        img = img.convert("RGB")
        img.thumbnail((THUMB_PX, THUMB_PX))
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=70)
        buf.seek(0)
        return buf
    except Exception as e:
        logger.warning("Could not decode image for export thumbnail: %s", e)
        return None


def _row_max_images(worker: dict) -> int:
    return max((len(worker.get(field) or []) for _, field in IMAGE_COLUMNS), default=0)


def build_workers_xlsx(workers: List[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Workers"

    n_text_cols = len(TEXT_COLUMNS)

    # Header row
    for i, (label, _) in enumerate(TEXT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(label) + 2)

    image_col_start = n_text_cols + 1
    for j, (label, _) in enumerate(IMAGE_COLUMNS):
        col = image_col_start + j
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        # Width sized to fit several thumbnails side-by-side horizontally is
        # impractical in one cell; instead we stack vertically and widen once.
        ws.column_dimensions[get_column_letter(col)].width = (THUMB_PX + 8) * COL_WIDTH_PX_TO_CHARS * 7

    row_idx = 2
    for w in workers:
        for i, (label, field) in enumerate(TEXT_COLUMNS, start=1):
            if label == "Languages":
                value = ", ".join(w.get("languages") or [])
            elif label == "Skills":
                value = ", ".join(w.get("skills") or [])
            elif label == "Duplicate Flags":
                value = " | ".join(w.get("duplicate_flags") or [])
            else:
                value = w.get(field)
            ws.cell(row=row_idx, column=i, value=value if value not in (None, "") else "")

        max_imgs = _row_max_images(w)
        row_height_px = max(THUMB_PX, max_imgs * (THUMB_PX + 6)) if max_imgs else 20
        ws.row_dimensions[row_idx].height = row_height_px * ROW_HEIGHT_PX_TO_PT

        for j, (_, field) in enumerate(IMAGE_COLUMNS):
            col = image_col_start + j
            images = w.get(field) or []
            for k, data_url in enumerate(images):
                thumb = _decode_thumbnail(data_url)
                if not thumb:
                    continue
                xl_img = XLImage(thumb)
                xl_img.width = THUMB_PX
                xl_img.height = THUMB_PX
                # Anchor each image at the top-left of its cell, offset
                # downward per index within the same field via pixel offset.
                marker = AnchorMarker(col=col - 1, colOff=pixels_to_EMU(4), row=row_idx - 1, rowOff=pixels_to_EMU(4 + k * (THUMB_PX + 6)))
                size = XDRPositiveSize2D(pixels_to_EMU(THUMB_PX), pixels_to_EMU(THUMB_PX))
                xl_img.anchor = OneCellAnchor(_from=marker, ext=size)
                ws.add_image(xl_img)

        row_idx += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
